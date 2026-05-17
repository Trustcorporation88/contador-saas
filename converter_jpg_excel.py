#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para converter dados de extratos de cartão de crédito (JPG) para Excel (XLS)
Usa OCR para extrair dados das imagens
"""

import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Tentar usar pytesseract com Tesseract
reader = None
try:
    import pytesseract
    from PIL import Image
    print("Tentando usar Tesseract OCR...")
    reader = 'tesseract'
except ImportError:
    print("Pytesseract não disponível, tentando EasyOCR...")
    try:
        import easyocr
        print("Inicializando OCR com EasyOCR (pode levar alguns segundos)...")
        reader = easyocr.Reader(['pt'], gpu=False)
    except ImportError:
        print("✗ Nenhuma library de OCR disponível")
        print("Por favor, instale um dos seguintes:")
        print("  - pytesseract + Tesseract OCR (https://github.com/UB-Mannheim/tesseract/wiki)")
        print("  - ou execute: pip install easyocr")
        exit(1)

def extract_text_from_image(image_path):
    """Extrai texto de uma imagem usando OCR disponível"""
    print(f"Processando: {image_path}")
    
    if reader == 'tesseract':
        try:
            from PIL import Image
            img = Image.open(image_path)
            # Tenta português, se falhar usa inglês
            try:
                text = pytesseract.image_to_string(img, lang='por')
            except:
                text = pytesseract.image_to_string(img, lang='eng')
            return text
        except Exception as e:
            print(f"  Erro com Tesseract: {e}")
            # Fallback: tenta com inglês sem language model
            try:
                from PIL import Image
                img = Image.open(image_path)
                text = pytesseract.image_to_string(img)
                return text
            except:
                return ""
    else:
        # EasyOCR
        try:
            result = reader.readtext(str(image_path), detail=0)
            return '\n'.join(result)
        except Exception as e:
            print(f"  Erro com EasyOCR: {e}")
            return ""

def parse_transaction_data(text):
    """Parse dos dados de transações do extrato"""
    transactions = []
    lines = text.split('\n')
    
    current_transaction = {}
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Padrão: Data (DD/MM) no início da linha
        date_pattern = r'^(\d{2}/\d{2})'
        date_match = re.match(date_pattern, line)
        
        if date_match:
            # Se há uma transação anterior, adiciona à lista
            if current_transaction:
                transactions.append(current_transaction)
            
            # Inicia nova transação
            current_transaction = {'Data': date_match.group(1)}
            rest_of_line = line[len(date_match.group(1)):].strip()
            
            # Tenta extrair valor em R$ (geralmente está no final)
            rs_pattern = r'(-?\d+[.,]\d{2})$'
            rs_match = re.search(rs_pattern, rest_of_line)
            
            if rs_match:
                current_transaction['R$'] = rs_match.group(1)
                current_transaction['Estabelecimento'] = rest_of_line[:rs_match.start()].strip()
            else:
                current_transaction['Estabelecimento'] = rest_of_line
        
        elif current_transaction and line and not re.match(r'^(Localidade|Moeda|Valor|Data|continua)', line):
            # Adiciona informações adicionais à transação
            if 'Localidade' not in current_transaction and any(c.isupper() for c in line):
                current_transaction['Localidade'] = line
    
    # Adiciona última transação
    if current_transaction:
        transactions.append(current_transaction)
    
    return transactions

def process_jpg_files(folder_path):
    """Processa todos os arquivos JPG da pasta"""
    all_transactions = []
    jpg_files = list(Path(folder_path).glob('*.jpg'))
    
    print(f"Encontrados {len(jpg_files)} arquivos JPG")
    
    for jpg_file in sorted(jpg_files):
        try:
            # Extrai texto da imagem
            text = extract_text_from_image(jpg_file)
            
            # Parse dos dados
            transactions = parse_transaction_data(text)
            all_transactions.extend(transactions)
            print(f"  ✓ {len(transactions)} transações extraídas")
        except Exception as e:
            print(f"  ✗ Erro ao processar {jpg_file}: {e}")
    
    return all_transactions

def create_excel_file(transactions, output_path):
    """Cria arquivo Excel com os dados extraídos"""
    # Criar DataFrame
    df = pd.DataFrame(transactions)
    
    # Reordenar colunas
    desired_columns = ['Data', 'Estabelecimento', 'Localidade', 'Moeda Origem', 
                       'Valor Origem', 'US$', 'Cotação', 'R$']
    existing_columns = [col for col in desired_columns if col in df.columns]
    df = df[existing_columns]
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transações"
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escrever dados
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Formatar cabeçalho
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Ajustar largura das colunas
    widths = {
        'A': 12,  # Data
        'B': 30,  # Estabelecimento
        'C': 20,  # Localidade
        'D': 15,  # Moeda Origem
        'E': 15,  # Valor Origem
        'F': 12,  # US$
        'G': 12,  # Cotação
        'H': 15   # R$
    }
    
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Salvar arquivo
    wb.save(output_path)
    print(f"\n✓ Arquivo Excel criado: {output_path}")
    print(f"  Total de transações: {len(df)}")

def main():
    folder_path = r"c:\jpg"
    output_path = r"c:\jpg\Extratos_Cartao_Credito.xlsx"
    
    print("=" * 60)
    print("CONVERTER JPG PARA EXCEL - EXTRATOS DE CARTÃO")
    print("=" * 60)
    
    # Processar arquivos JPG
    transactions = process_jpg_files(folder_path)
    
    if transactions:
        # Criar arquivo Excel
        create_excel_file(transactions, output_path)
        print("\n✓ Processo concluído com sucesso!")
    else:
        print("\n✗ Nenhuma transação foi extraída")

if __name__ == "__main__":
    main()
