#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para converter dados de extratos de cartão de crédito para Excel
Abordagem: Estruturação dos dados visualizados nas imagens
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Dados extraídos manualmente das imagens JPG
# Estrutura: Data, Estabelecimento, Localidade, Moeda Origem, Valor Origem, US$, Cotação, R$

transactions_data = [
    # Cartão CYNTIA D O M RINALDI (473270*****1067) - Página 05/12
    {'Data': '16/03', 'Estabelecimento': 'MERCAD*ACQUEPARC02/06', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 30.45},
    {'Data': '18/03', 'Estabelecimento': 'LOCITANE BUEUPARCO2/03', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 77.56},
    {'Data': '18/03', 'Estabelecimento': 'MP*BANCORRADESPARCO2/02', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 106.43},
    {'Data': '18/03', 'Estabelecimento': 'DROGARIA SP DRPARCO2/02', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 93.30},
    {'Data': '18/03', 'Estabelecimento': 'AREZZO PARCO2/03', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 133.30},
    {'Data': '19/03', 'Estabelecimento': 'ZP*GAVEA NEGOPARCO2/12', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 1474.16},
    {'Data': '20/03', 'Estabelecimento': 'ZIG PARCO2/02', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 110.00},
    {'Data': '21/03', 'Estabelecimento': 'SEPHORA CATARIP PARCO2/06', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 45.71},
    {'Data': '21/03', 'Estabelecimento': 'AREZZO SHOPPINGPARCO2/06', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 133.60},
    {'Data': '26/03', 'Estabelecimento': 'DBM DAHYANA PARCO2/03', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 235.00},
    {'Data': '26/03', 'Estabelecimento': 'SORVEMIX PARCO2/03', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 140.38},
    {'Data': '27/03', 'Estabelecimento': 'ATACADAO 139 APARCO2/02', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 357.49},
    {'Data': '29/03', 'Estabelecimento': 'MP*FLAVIOLUSIRRPARCO2/02', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 56.87},
    {'Data': '30/03', 'Estabelecimento': 'ZP*GAVEA NEGOPARCO2/10', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 135.00},
    {'Data': '30/03', 'Estabelecimento': 'MERCADOLIVRE*MPARCO2/03', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 20.43},
    {'Data': '31/03', 'Estabelecimento': 'MP*APP PARCO2/03', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 27.42},
    {'Data': '31/03', 'Estabelecimento': 'MP*BANCOSANTANPARCO2/02', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 457.63},
    {'Data': '02/04', 'Estabelecimento': 'IFD*PET CENTERPARCO2/02', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 43.20},
    {'Data': '07/04', 'Estabelecimento': 'ASAA*GLAMOUR PARCO2/03', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 54.53},
    {'Data': '09/04', 'Estabelecimento': '139 ATACADAO OPARCO1/02', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 130.67},
    {'Data': '10/04', 'Estabelecimento': 'CASCA DE NOZ', 'Localidade': 'BAURU', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 64.80},
    {'Data': '10/04', 'Estabelecimento': 'CONFIANCA MAX', 'Localidade': 'BAURU', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 117.82},
    {'Data': '10/04', 'Estabelecimento': 'EBN*SPOTIFY', 'Localidade': 'CURITIBA', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 23.90},
    {'Data': '11/04', 'Estabelecimento': 'PM *CRAZY TILE', 'Localidade': 'HONGKONG', 'Moeda': 'BRL', 'Valor_Origem': 2.49, 'US': 0.51, 'Cotacao': 5.3466, 'R': 2.72},
    {'Data': '11/04', 'Estabelecimento': 'SHOPEE *MAGAZPARC01/03', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 58.28},
    {'Data': '11/04', 'Estabelecimento': 'PM *CRAZY TILE', 'Localidade': 'HONGKONG', 'Moeda': 'BRL', 'Valor_Origem': 4.99, 'US': 1.02, 'Cotacao': 5.3466, 'R': 5.45},
    {'Data': '11/04', 'Estabelecimento': 'PM *CRAZY TILE', 'Localidade': 'HONGKONG', 'Moeda': 'BRL', 'Valor_Origem': 4.99, 'US': 1.02, 'Cotacao': 5.3466, 'R': 5.45},
    {'Data': '12/04', 'Estabelecimento': 'MERCADOLIVRE*MEDICADOLIVRE', 'Localidade': 'MANAUA', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 24.89},
    {'Data': '12/04', 'Estabelecimento': 'PM *CRAZY TILE', 'Localidade': 'HONGKONG', 'Moeda': 'BRL', 'Valor_Origem': 9.99, 'US': 2.04, 'Cotacao': 5.3466, 'R': 10.90},
    {'Data': '12/04', 'Estabelecimento': 'PM *CRAZY TILE', 'Localidade': 'HONGKONG', 'Moeda': 'BRL', 'Valor_Origem': 9.99, 'US': 2.04, 'Cotacao': 5.3466, 'R': 10.90},
    {'Data': '13/04', 'Estabelecimento': 'CONTA VIVO', 'Localidade': 'SAO PAULO', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 385.85},
    {'Data': '13/04', 'Estabelecimento': 'APPLE.COM/BILL', 'Localidade': 'SAO PAULO', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 19.90},
    {'Data': '13/04', 'Estabelecimento': 'CENTRO DE VACINAS BAU', 'Localidade': 'BAURU', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 310.00},
    {'Data': '14/04', 'Estabelecimento': 'PADARIA COPACABANA', 'Localidade': 'BAURU', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 83.47},
]

# Adicionar mais transações do segundo cartão (FLAVIO L RINALDI) - Página 09/12
flavio_transactions = [
    {'Data': '14/12', 'Estabelecimento': 'AIRBNB *HKCH315/06', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 124.60},
    {'Data': '23/12', 'Estabelecimento': 'ESPECIALISTA GRACOS/06', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 90.00},
    {'Data': '26/12', 'Estabelecimento': 'PB*KASPERSKY BPARC05/06', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 18.31},
    {'Data': '13/01', 'Estabelecimento': 'HTM*ROBISON COSPARCO4/09', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 9.07},
    {'Data': '15/01', 'Estabelecimento': 'TICKAYWORKSHOPPARCO4/04', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 60.15},
    {'Data': '16/01', 'Estabelecimento': 'CFPL PAULISTA PARCO4/10', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 140.87},
    {'Data': '17/01', 'Estabelecimento': 'HTM*LISTA SECPARCO4/04', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 13.10},
    {'Data': '17/01', 'Estabelecimento': 'HTM*WORKSHOP -PARCO4/04', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 9.48},
    {'Data': '17/01', 'Estabelecimento': 'HTM*GRAVACAO -PARCO4/08', 'Localidade': '', 'Moeda': '', 'Valor_Origem': '', 'US': '', 'Cotacao': '', 'R': 14.70},
]

transactions_data.extend(flavio_transactions)

def create_excel_file(data, output_path):
    """Cria arquivo Excel formatado com os dados"""
    
    df = pd.DataFrame(data)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transações Cartão"
    
    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Colunas
    columns = ['Data', 'Estabelecimento', 'Localidade', 'Moeda', 'Valor_Origem', 'US', 'Cotacao', 'R']
    
    # Escrever cabeçalho
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Escrever dados
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Formatação
            if col_name == 'R':
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            elif col_name in ['Valor_Origem', 'US', 'Cotacao']:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Ajustar largura das colunas
    widths = {
        'A': 10,   # Data
        'B': 35,   # Estabelecimento
        'C': 18,   # Localidade
        'D': 12,   # Moeda
        'E': 14,   # Valor_Origem
        'F': 12,   # US
        'G': 12,   # Cotacao
        'H': 14    # R
    }
    
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    # Salvar
    wb.save(output_path)
    print(f"✓ Arquivo Excel criado com sucesso: {output_path}")
    print(f"  Total de transações: {len(data)}")
    print(f"  Intervalo: {data[0]['Data']} a {data[-1]['Data']}")

def main():
    output_path = r"c:\jpg\Extratos_Cartao_Credito.xlsx"
    
    print("=" * 70)
    print("CONVERTER EXTRATOS DE CARTÃO (JPG) PARA EXCEL (XLS)")
    print("=" * 70)
    print(f"\nTotalizando {len(transactions_data)} transações...")
    
    create_excel_file(transactions_data, output_path)
    
    print("\n✓ Processo concluído com sucesso!")
    print(f"\nArquivo: {output_path}")

if __name__ == "__main__":
    main()
